"""Кастомные тулы агента: разведка Excel, бэкап, построчная LLM-обработка."""
from __future__ import annotations

import shutil
from pathlib import Path

from langchain_core.tools import tool


@tool
def inspect_excel(path: str) -> str:
    """Быстрая разведка Excel-файла: листы, размер, типы столбцов и первые 5 строк.

    Использовать ПЕРЕД любой работой с файлом вместо чтения его целиком.
    path — абсолютный путь к .xlsx.
    """
    import pandas as pd

    p = Path(path)
    if not p.exists():
        return f"Файл не найден: {p}"
    out: list[str] = []
    xl = pd.ExcelFile(p)
    out.append(f"Листы: {xl.sheet_names}")
    for sheet in xl.sheet_names:
        df = xl.parse(sheet)
        out.append(
            f"\n--- Лист '{sheet}': {df.shape[0]} строк x {df.shape[1]} столбцов"
        )
        out.append(f"Типы: {df.dtypes.to_dict()}")
        out.append(f"head(5):\n{df.head(5).to_string()}")
    return "\n".join(out)


@tool
def backup_file(path: str) -> str:
    """Создаёт резервную копию файла рядом с ним: <имя>.bak.

    Вызывать один раз ПЕРЕД первым изменением существующего файла.
    Если .bak уже есть — не перезаписывает (бэкап хранит исходное состояние).
    """
    p = Path(path)
    if not p.exists():
        return f"Файл не найден: {p}"
    bak = p.with_suffix(p.suffix + ".bak")
    if bak.exists():
        return f"Бэкап уже существует, не трогаю: {bak}"
    shutil.copy2(p, bak)
    return f"Бэкап создан: {bak}"


@tool
def llm_process_column(
    path: str,
    sheet: str,
    source_column: str,
    new_column: str,
    instruction: str,
    batch_size: int = 20,
) -> str:
    """Построчная обработка текстового столбца LLM-кой: каждая строка прогоняется
    через модель по инструкции, результат пишется в новый столбец.

    Использовать для задач вида «определи тональность каждого отзыва»,
    «классифицируй каждую строку», «извлеки из каждой ячейки X» — когда нужен
    проход по ВСЕМ строкам, а не общая оценка. Строки обрабатываются циклом
    пакетами, таблица в контекст агента не попадает.

    path — абсолютный путь к .xlsx; sheet — имя листа;
    source_column — заголовок столбца с текстом; new_column — заголовок нового
    столбца с результатами; instruction — что сделать с каждым значением
    (например: «Определи тональность отзыва. Ответь одним словом: позитивный
    или негативный»); batch_size — строк на один запрос к модели.
    """
    from openpyxl import load_workbook

    from .config import make_map_model

    p = Path(path)
    if not p.exists():
        return f"Файл не найден: {p}"
    wb = load_workbook(p)
    if sheet not in wb.sheetnames:
        return f"Нет листа '{sheet}'. Есть: {wb.sheetnames}"
    ws = wb[sheet]

    headers = {str(ws.cell(row=1, column=c).value): c for c in range(1, ws.max_column + 1)}
    if source_column not in headers:
        return f"Нет столбца '{source_column}'. Есть: {list(headers)}"
    src_col = headers[source_column]
    dst_col = headers.get(new_column) or ws.max_column + 1

    values = [
        (r, ws.cell(row=r, column=src_col).value)
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=src_col).value is not None
    ]
    if not values:
        return f"Столбец '{source_column}' пуст."

    model = make_map_model()
    results: dict[int, str] = {}
    failed: list[int] = []
    for start in range(0, len(values), batch_size):
        batch = values[start:start + batch_size]
        numbered = "\n".join(f"{i + 1}. {str(v)[:500]}" for i, (_, v) in enumerate(batch))
        prompt = (
            f"{instruction}\n\nОбработай КАЖДЫЙ пункт списка. Ответь строго в формате "
            f"'<номер>. <ответ>' — по одной строке на пункт, без пояснений.\n\n{numbered}"
        )
        try:
            text = model.invoke(prompt).content
        except Exception as e:  # noqa: BLE001
            failed.extend(r for r, _ in batch)
            results.update({r: f"ОШИБКА: {e}" for r, _ in batch})
            continue
        parsed: dict[int, str] = {}
        for line in str(text).splitlines():
            line = line.strip()
            if "." in line and line.split(".", 1)[0].strip().isdigit():
                n, ans = line.split(".", 1)
                parsed[int(n.strip())] = ans.strip()
        for i, (r, _) in enumerate(batch):
            if i + 1 in parsed:
                results[r] = parsed[i + 1]
            else:
                failed.append(r)
                results[r] = "?"

    ws.cell(row=1, column=dst_col, value=new_column)
    for r, ans in results.items():
        ws.cell(row=r, column=dst_col, value=ans)
    wb.save(p)

    msg = (f"Обработано {len(results)} строк из '{source_column}' -> '{new_column}' "
           f"({(len(values) - 1) // batch_size + 1} запросов к LLM).")
    if failed:
        msg += f" Не распознано/ошибки: {len(failed)} строк, помечены '?'."
    return msg


CUSTOM_TOOLS = [inspect_excel, backup_file, llm_process_column]
