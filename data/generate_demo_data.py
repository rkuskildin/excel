"""Генерация демо-данных: «грязные» эксельки консалтингового кейса.

Запуск:  python data/generate_demo_data.py
Создаёт файлы в data/source/ (эталон, не изменяется агентом).
"""
from __future__ import annotations

import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "source"
OUT.mkdir(parents=True, exist_ok=True)
random.seed(42)

REGIONS = ["Москва", "СПб", "Казань", "Екатеринбург", "Новосибирск"]
PRODUCTS = ["Аудит", "Налоговый консалтинг", "Стратегия", "Due Diligence", "IT-консалтинг"]
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def style_header(ws, widths):
    for col, width in enumerate(widths, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(col)].width = width


def dirty_money(value: float) -> str:
    """Превращает число в «битую» строку: '1 234,50', '1 234,50 руб.', '1234.5'."""
    style = random.choice(["ru", "plain", "spaces", "rub"])
    ru = f"{value:,.2f}".replace(",", " ").replace(".", ",")
    if style == "ru":
        return ru
    if style == "spaces":
        return f"  {value:,.0f}  ".replace(",", " ")
    if style == "rub":
        return f"{ru} руб."
    return f"{value:.2f}"


def make_sales():
    """sales_2025.xlsx — продажи с «битым» столбцом выручки и дублями."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Продажи"
    headers = ["ID", "Дата", "Регион", "Услуга", "Часы", "Ставка", "Выручка (грязная)"]
    ws.append(headers)
    rows = []
    for i in range(1, 61):
        hours = random.randint(8, 120)
        rate = random.choice([3500, 5000, 7500, 12000])
        rows.append([
            i,
            f"2025-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}",
            random.choice(REGIONS),
            random.choice(PRODUCTS),
            hours,
            rate,
            dirty_money(hours * rate),
        ])
    # дубли: 5 строк повторяются
    rows += [list(r) for r in random.sample(rows, 5)]
    random.shuffle(rows)
    for r in rows:
        ws.append(r)
    style_header(ws, [6, 12, 16, 22, 8, 10, 20])
    wb.save(OUT / "sales_2025.xlsx")


def make_attendance():
    """attendance_q1/q2.xlsx — два списка посещаемости с пересечениями (кейс из лекции)."""
    names = [
        "Иванов И.И.", "Петров П.П.", "Сидорова А.А.", "Кузнецов К.К.",
        "Смирнова Е.Е.", "Волков В.В.", "Зайцева З.З.", "Морозов М.М.",
        "Павлова П.А.", "Соколов С.С.", "Лебедева Л.Л.", "Козлов О.О.",
    ]
    for fname, sample in [
        ("attendance_q1.xlsx", names[:8]),
        ("attendance_q2.xlsx", names[4:]),
    ]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Посещаемость"
        ws.append(["ФИО", "Группа"])
        for n in sample:
            ws.append([n, random.choice(["А-101", "А-102"])])
        style_header(ws, [25, 10])
        wb.save(OUT / fname)


def make_clients():
    """clients.xlsx — справочник клиентов с пропусками в столбце «Сегмент»."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    ws.append(["Клиент", "Регион", "Выручка клиента, млн", "Сегмент"])
    for i in range(1, 31):
        rev = round(random.uniform(5, 900), 1)
        segment = None if random.random() < 0.4 else (
            "Крупный" if rev > 300 else "Средний" if rev > 50 else "Малый")
        ws.append([f"Клиент-{i:02d}", random.choice(REGIONS), rev, segment])
    style_header(ws, [14, 16, 22, 12])
    wb.save(OUT / "clients.xlsx")


# Отзывы с известной разметкой — эталон для задачи построчной LLM-обработки.
REVIEWS = [
    ("Отличная работа консультантов, отчёт превзошёл ожидания!", "позитивный"),
    ("Сорваны все сроки, отчёт пришлось переделывать самим.", "негативный"),
    ("Очень довольны сотрудничеством, рекомендуем коллегам.", "позитивный"),
    ("Качество анализа ужасное, данные перепутаны.", "негативный"),
    ("Команда быстро вникла в специфику, всё чётко и по делу.", "позитивный"),
    ("За такие деньги ожидали большего, результат разочаровал.", "негативный"),
    ("Прекрасная презентация результатов, руководство в восторге.", "позитивный"),
    ("Менеджер не отвечал на письма неделями, кошмар.", "негативный"),
    ("Помогли сэкономить десятки часов ручной работы, спасибо!", "позитивный"),
    ("Модель построена с грубыми ошибками, доверия ноль.", "негативный"),
    ("Глубокая экспертиза в рисках, продлеваем контракт.", "позитивный"),
    ("Скучные шаблонные выводы, никакой пользы для бизнеса.", "негативный"),
    ("Сильная команда, отличные рекомендации по оптимизации.", "позитивный"),
    ("Постоянные переносы встреч, проект затянулся вдвое.", "негативный"),
    ("Лучший подрядчик из всех, с кем работали. Браво!", "позитивный"),
    ("Итоговый файл был битый и не открывался, ужасно.", "негативный"),
    ("Аккуратные таблицы, всё сходится до копейки, молодцы.", "позитивный"),
    ("Передали задачу стажёрам, уровень работы плачевный.", "негативный"),
    ("Оперативно исправили все замечания, приятно работать.", "позитивный"),
    ("Завысили смету в два раза без объяснений, обман.", "негативный"),
]


def make_reviews():
    """reviews.xlsx — отзывы клиентов для построчной LLM-классификации."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Отзывы"
    ws.append(["ID", "Отзыв"])
    for i, (text, _) in enumerate(REVIEWS, 1):
        ws.append([i, text])
    style_header(ws, [6, 70])
    wb.save(OUT / "reviews.xlsx")


if __name__ == "__main__":
    make_sales()
    make_attendance()
    make_clients()
    make_reviews()
    print(f"OK: данные созданы в {OUT}")
