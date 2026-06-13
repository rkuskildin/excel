"""Детерминированные валидаторы задач мини-бенча.

Каждый валидатор: validate(workdir: Path) -> (passed: bool, message: str).
Эталон считается от data/source (исходных, неизменённых файлов).
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "source"


def _rows_as_set(df: pd.DataFrame) -> set:
    return {tuple(str(v) for v in row) for row in df.itertuples(index=False)}


def t01_dedup(workdir: Path):
    src = pd.read_excel(SOURCE / "sales_2025.xlsx", sheet_name="Продажи")
    expected = src.drop_duplicates()
    got = pd.read_excel(workdir / "sales_2025.xlsx", sheet_name="Продажи")
    if len(got) != len(expected):
        return False, f"строк {len(got)}, ожидалось {len(expected)}"
    if _rows_as_set(got) != _rows_as_set(expected):
        return False, "набор строк не совпадает с эталоном"
    return True, f"OK: {len(got)} уникальных строк"


def _parse_dirty(x) -> float:
    if isinstance(x, (int, float)):
        return float(x)
    s = (str(x).replace("\xa0", " ").replace("руб.", "").replace("руб", "")
         .replace(" ", "").replace(",", "."))
    return float(s.strip(" ."))


def t02_clean_revenue(workdir: Path):
    got = pd.read_excel(workdir / "sales_2025.xlsx", sheet_name="Продажи")
    src = pd.read_excel(SOURCE / "sales_2025.xlsx", sheet_name="Продажи")
    if "Выручка" not in got.columns:
        return False, "нет столбца 'Выручка'"
    if len(got) != len(src):
        return False, f"строк {len(got)}, ожидалось {len(src)} (строки удалять нельзя)"
    bad = 0
    for _, row in got.iterrows():
        try:
            v = float(row["Выручка"])
        except (TypeError, ValueError):
            bad += 1
            continue
        if abs(v - row["Часы"] * row["Ставка"]) > 0.01:
            bad += 1
    if bad:
        return False, f"{bad} строк с неверной выручкой"
    return True, f"OK: {len(got)} строк, сумма {got['Выручка'].sum():,.0f}"


def _check_svod(f: Path):
    """Общая проверка свода по регионам для T03/T07."""
    sales = pd.read_excel(f, sheet_name="Продажи")
    src = pd.read_excel(SOURCE / "sales_2025.xlsx", sheet_name="Продажи")
    if len(sales) != len(src):
        return False, "исходный лист 'Продажи' изменён"
    expected = (sales["Часы"] * sales["Ставка"]).groupby(sales["Регион"]).sum().to_dict()
    svod = pd.read_excel(f, sheet_name="Свод")
    cols = {str(c).strip(): c for c in svod.columns}
    if "Регион" not in cols or "Выручка" not in cols:
        return False, f"ожидались столбцы 'Регион' и 'Выручка', есть {list(svod.columns)}"
    got = dict(zip(svod[cols["Регион"]], svod[cols["Выручка"]]))
    for region, total in expected.items():
        if region not in got:
            return False, f"нет региона {region}"
        if abs(float(got[region]) - total) > 0.01:
            return False, f"{region}: {got[region]} != {total}"
    return True, f"OK: {len(expected)} регионов сходятся"


def t03_region_summary(workdir: Path):
    f = workdir / "sales_2025.xlsx"
    wb = load_workbook(f, read_only=True)
    has = "Свод" in wb.sheetnames
    wb.close()
    if not has:
        return False, "нет листа 'Свод'"
    return _check_svod(f)


def t04_merge_attendance(workdir: Path):
    f = workdir / "attendance_all.xlsx"
    if not f.exists():
        return False, "нет файла attendance_all.xlsx"
    q1 = pd.read_excel(SOURCE / "attendance_q1.xlsx")
    q2 = pd.read_excel(SOURCE / "attendance_q2.xlsx")
    expected = set(q1["ФИО"]) | set(q2["ФИО"])
    got = pd.read_excel(f)
    if "ФИО" not in got.columns:
        return False, f"нет столбца 'ФИО', есть {list(got.columns)}"
    names = list(got["ФИО"])
    if len(names) != len(set(names)):
        return False, "есть дубли по ФИО"
    if set(names) != expected:
        return False, f"состав не совпадает: {set(names) ^ expected}"
    return True, f"OK: {len(names)} уникальных человек"


def t05_fill_segment(workdir: Path):
    got = pd.read_excel(workdir / "clients.xlsx")
    src = pd.read_excel(SOURCE / "clients.xlsx")
    if len(got) != len(src):
        return False, "число строк изменилось"
    rev_col = "Выручка клиента, млн"
    for i in range(len(src)):
        rev = src.loc[i, rev_col]
        rule = "Крупный" if rev > 300 else "Средний" if rev > 50 else "Малый"
        old, new = src.loc[i, "Сегмент"], got.loc[i, "Сегмент"]
        if pd.isna(new):
            return False, f"строка {i + 2}: сегмент не заполнен"
        if pd.notna(old) and new != old:
            return False, f"строка {i + 2}: изменено существующее значение"
        if pd.isna(old) and new != rule:
            return False, f"строка {i + 2}: '{new}', по правилу '{rule}' (выручка {rev})"
    return True, "OK: все пропуски заполнены по правилу, старые значения целы"


def t06_rename_keep_format(workdir: Path):
    wb = load_workbook(workdir / "clients.xlsx")
    ws = wb["Клиенты"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    if headers[2] != "Выручка, млн руб.":
        return False, f"заголовок не переименован: {headers[2]!r}"
    for c in range(1, 5):
        cell = ws.cell(row=1, column=c)
        fg = cell.fill.fgColor.rgb or ""
        if not str(fg).endswith("4472C4"):
            return False, f"потеряна заливка заголовка в столбце {c} (fg={fg})"
        if not cell.font.bold:
            return False, f"потерян жирный шрифт заголовка в столбце {c}"
    expected_widths = [14, 16, 22, 12]
    for i, w in enumerate(expected_widths, 1):
        got_w = ws.column_dimensions[chr(64 + i)].width
        if got_w is None or abs(got_w - w) > 1:
            return False, f"ширина столбца {chr(64 + i)}: {got_w}, ожидалось ~{w}"
    return True, "OK: заголовок переименован, форматирование сохранено"


def t07_native_chart(workdir: Path):
    f = workdir / "sales_2025.xlsx"
    wb = load_workbook(f)  # не read_only: нужны объекты диаграмм
    if "Свод" not in wb.sheetnames:
        return False, "нет листа 'Свод'"
    charts = wb["Свод"]._charts
    if not charts:
        return False, "на листе 'Свод' нет нативной диаграммы Excel"
    kind = type(charts[0]).__name__
    ok, msg = _check_svod(f)
    if not ok:
        return False, f"диаграмма есть, но {msg}"
    return True, f"OK: диаграмма {kind}, {msg}"


def t08_review_sentiment(workdir: Path, threshold: float = 0.9):
    sys.path.insert(0, str(ROOT / "data"))
    from generate_demo_data import REVIEWS  # эталонная разметка

    got = pd.read_excel(workdir / "reviews.xlsx", sheet_name="Отзывы")
    if "Тональность" not in got.columns:
        return False, "нет столбца 'Тональность'"
    if len(got) != len(REVIEWS):
        return False, f"строк {len(got)}, ожидалось {len(REVIEWS)}"
    truth = {text: label for text, label in REVIEWS}
    correct = 0
    for _, row in got.iterrows():
        label = str(row["Тональность"]).strip().lower().rstrip(".")
        if label == truth.get(str(row["Отзыв"]), "?"):
            correct += 1
    acc = correct / len(REVIEWS)
    if acc < threshold:
        return False, f"точность {correct}/{len(REVIEWS)} ({acc:.0%}) < порога {threshold:.0%}"
    return True, f"OK: точность {correct}/{len(REVIEWS)} ({acc:.0%})"


VALIDATORS = {
    "T01_dedup": t01_dedup,
    "T02_clean_revenue": t02_clean_revenue,
    "T03_region_summary": t03_region_summary,
    "T04_merge_attendance": t04_merge_attendance,
    "T05_fill_segment": t05_fill_segment,
    "T06_rename_keep_format": t06_rename_keep_format,
    "T07_native_chart": t07_native_chart,
    "T08_review_sentiment": t08_review_sentiment,
}

if __name__ == "__main__":
    # самопроверка валидатора: python bench/validators.py <task_id> <workdir>
    task_id, wd = sys.argv[1], Path(sys.argv[2])
    ok, msg = VALIDATORS[task_id](wd)
    print(("PASS" if ok else "FAIL"), msg)
